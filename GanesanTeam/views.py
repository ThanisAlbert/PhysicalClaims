from datetime import datetime
from io import BytesIO

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.timezone import make_naive
from openpyxl.reader.excel import load_workbook
from tablib import Dataset

from GanesanTeam.admin import claimtrackerresource
from GanesanTeam.forms import UploadFileForm
from GanesanTeam.models import Login, Department, Divisions, ClaimType, Regions, Vendors_Brands, Category, \
    Forwarder_Code, ClaimsStatus, AdditionalClaimsStatus, Status, ClaimsTracker, Currency, Serilized_NonSerilized
import logging
import pandas as pd
from django.http import HttpResponse
from .models import ClaimsTracker
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
logger = logging.getLogger('django')
from .models import ClaimsTracker

def export_claims_status(request):

    claimstatusqueue = request.POST["Claimstatusqueue"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Claims Tracker"

    # Define the column headers
    columns = [
        'id', 'Department', 'Divisions', 'ClaimType', 'Regions', 'RequestFrom',
        'Originator', 'ClaimedDate', 'ClaimRefNumber', 'Ponumber', 'cbn', 'hporder',
        'customer_request_date', 'Partner_Customername', 'CustomerCode',
        'Partner_location', 'CustomerReference', 'RedingtonInvoice',
        'Redington_Invoice_Date', 'SellUnitPrice', 'TotalSellPrice', 'Currency',
        'Customer_prf_Number', 'Vendors_Brands', 'Category', 'Partnumber', 'Qty',
        'PackId', 'BoxId', 'Serilized_NonSerilized', 'SerialNumber', 'UnitPrice',
        'TotalAmount', 'VendorInvNo', 'ECI_Date', 'DueDate_ExpiryDate', 'ASP_FDR_Date',
        'Call_Id', 'Detailed_Issue_Product', 'PRF_Num', 'RR_Num', 'RR_Date',
        'Unit_Collection_Date', 'CM_Num', 'CM_Amount', 'RPO_Num', 'RPO_Date',
        'GRN_Date', 'DHL_AWB_Num', 'Forwarder_Code', 'Claims_Status',
        'AdditionalClaimsstatus', 'Remarks', 'Status', 'CN_No', 'CN_Date', 'CN_Value',
        'CA', 'Ageingdays'
    ]

    # Write the column headers to the first row
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Create a date style
    date_style = NamedStyle(name="date_style", number_format="DD-MM-YYYY")

    # Query the database and write the data rows
    for row_num, claim in enumerate(ClaimsTracker.objects.filter(Claims_Status=claimstatusqueue), 2):
        ws.cell(row=row_num, column=1, value=claim.id)
        ws.cell(row=row_num, column=2, value=claim.Department)
        ws.cell(row=row_num, column=3, value=claim.Divisions)
        ws.cell(row=row_num, column=4, value=claim.ClaimType)
        ws.cell(row=row_num, column=5, value=claim.Regions)
        ws.cell(row=row_num, column=6, value=claim.RequestFrom)
        ws.cell(row=row_num, column=7, value=claim.Originator)
        ws.cell(row=row_num, column=8, value=make_naive(
            claim.ClaimedDate) if claim.ClaimedDate else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=9, value=claim.ClaimRefNumber)
        ws.cell(row=row_num, column=10, value=claim.Ponumber)
        ws.cell(row=row_num, column=11, value=claim.cbn)
        ws.cell(row=row_num, column=12, value=claim.hporder)
        ws.cell(row=row_num, column=13, value=make_naive(
            claim.customer_request_date) if claim.customer_request_date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=14, value=claim.Partner_Customername)
        ws.cell(row=row_num, column=15, value=claim.CustomerCode)
        ws.cell(row=row_num, column=16, value=claim.Partner_location)
        ws.cell(row=row_num, column=17, value=claim.CustomerReference)
        ws.cell(row=row_num, column=18, value=claim.RedingtonInvoice)
        ws.cell(row=row_num, column=19, value=make_naive(
            claim.Redington_Invoice_Date) if claim.Redington_Invoice_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=20, value=claim.SellUnitPrice)
        ws.cell(row=row_num, column=21, value=claim.TotalSellPrice)
        ws.cell(row=row_num, column=22, value=claim.Currency)
        ws.cell(row=row_num, column=23, value=claim.Customer_prf_Number)
        ws.cell(row=row_num, column=24, value=claim.Vendors_Brands)
        ws.cell(row=row_num, column=25, value=claim.Category)
        ws.cell(row=row_num, column=26, value=claim.Partnumber)
        ws.cell(row=row_num, column=27, value=claim.Qty)
        ws.cell(row=row_num, column=28, value=claim.PackId)
        ws.cell(row=row_num, column=29, value=claim.BoxId)
        ws.cell(row=row_num, column=30, value=claim.Serilized_NonSerilized)
        ws.cell(row=row_num, column=31, value=claim.SerialNumber)
        ws.cell(row=row_num, column=32, value=claim.UnitPrice)
        ws.cell(row=row_num, column=33, value=claim.TotalAmount)
        ws.cell(row=row_num, column=34, value=claim.VendorInvNo)
        ws.cell(row=row_num, column=35,
                value=make_naive(claim.ECI_Date) if claim.ECI_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=36, value=make_naive(
            claim.DueDate_ExpiryDate) if claim.DueDate_ExpiryDate else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=37, value=make_naive(
            claim.ASP_FDR_Date) if claim.ASP_FDR_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=38, value=claim.Call_Id)
        ws.cell(row=row_num, column=39, value=claim.Detailed_Issue_Product)
        ws.cell(row=row_num, column=40, value=claim.PRF_Num)
        ws.cell(row=row_num, column=41, value=claim.RR_Num)
        ws.cell(row=row_num, column=42,
                value=make_naive(claim.RR_Date) if claim.RR_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=43, value=make_naive(
            claim.Unit_Collection_Date) if claim.Unit_Collection_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=44, value=claim.CM_Num)
        ws.cell(row=row_num, column=45, value=claim.CM_Amount)
        ws.cell(row=row_num, column=46, value=claim.RPO_Num)
        ws.cell(row=row_num, column=47,
                value=make_naive(claim.RPO_Date) if claim.RPO_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=48,
                value=make_naive(claim.GRN_Date) if claim.GRN_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=49, value=claim.DHL_AWB_Num)
        ws.cell(row=row_num, column=50, value=claim.Forwarder_Code)
        ws.cell(row=row_num, column=51, value=claim.Claims_Status)
        ws.cell(row=row_num, column=52, value=claim.AdditionalClaimsstatus)
        ws.cell(row=row_num, column=53, value=claim.Remarks)
        ws.cell(row=row_num, column=54, value=claim.Status)
        ws.cell(row=row_num, column=55, value=claim.CN_No)
        ws.cell(row=row_num, column=56,
                value=make_naive(claim.CN_Date) if claim.CN_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=57, value=claim.CN_Value)
        ws.cell(row=row_num, column=58, value=claim.CA)
        ws.cell(row=row_num, column=59, value=claim.Ageingdays)

    # Set the response content type and headers
    current_date = datetime.now().strftime('%d%m%Y')  # Format: YYYYMMDD
    filename = f'claims_tracker_{current_date}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save the workbook to the response
    wb.save(response)

    return response


def claimqueueview(request):

    claimstatusqueue = request.POST["Claimstatusqueue"]

    ClaimsTrackers = ClaimsTracker.objects.filter(Claims_Status=claimstatusqueue)

    Departments = Department.objects.all()
    Divisionss = Divisions.objects.all()
    ClaimTypes = ClaimType.objects.all()
    Regionss = Regions.objects.all()
    Vendors_Brandss = Vendors_Brands.objects.all()
    Categorys = Category.objects.all()
    ForwarderCodes = Forwarder_Code.objects.all()

    ClaimsStatuss = ClaimsStatus.objects.all()

    AdditionalClaimsStatuss = AdditionalClaimsStatus.objects.all()
    Statuss = Status.objects.all()

    Currencys = Currency.objects.all()
    Serilizeds = Serilized_NonSerilized.objects.all()

    Context = {'claimstatusqueue':claimstatusqueue,'Serilizeds': Serilizeds, 'Currencys': Currencys, 'username': request.session['username'],
               'ClaimsTrackers': ClaimsTrackers, 'Statuss': Statuss, 'AdditionalClaimsStatuss': AdditionalClaimsStatuss,
               'ClaimsStatuss': ClaimsStatuss, 'ForwarderCodes': ForwarderCodes, 'Categorys': Categorys,
               'Vendors_Brandss': Vendors_Brandss, 'Departments': Departments, 'Divisionss': Divisionss,
               'ClaimTypes': ClaimTypes, 'Regionss': Regionss}
    return render(request, 'index.html', context=Context)

def import_claims_tracker(request):
    if request.method == "POST":
        # Get the uploaded file
        excel_file = request.FILES['file']

        # Load the workbook and get the active sheet
        wb = load_workbook(filename=excel_file)
        ws = wb.active

        # Iterate over the rows in the sheet
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row

            id=row[0]

            if ClaimsTracker.objects.filter(id=id).exists():
                claimid = row[0]
            else:
                ids = ClaimsTracker.objects.values_list('id', flat=True)
                last_segments = [int(claim_id.split('-')[-1]) for claim_id in ids]
                max_value = max(last_segments) + 1
                max_value_formatted = f"{max_value:06}"
                claimid = str("REDCLMFY24-25-") + str(max_value_formatted)

            claim_data = {
                'id': claimid,
                'Department': row[1],
                'Divisions': row[2],
                'ClaimType': row[3],
                'Regions': row[4],
                'RequestFrom': row[5],
                'Originator': row[6],
                'ClaimedDate': row[7],
                'ClaimRefNumber': row[8],
                'Ponumber': row[9],
                'cbn': row[10],
                'hporder': row[11],
                'customer_request_date': row[12],
                'Partner_Customername': row[13],
                'CustomerCode': row[14],
                'Partner_location': row[15],
                'CustomerReference': row[16],
                'RedingtonInvoice': row[17],
                'Redington_Invoice_Date': row[18],
                'SellUnitPrice': row[19],
                'TotalSellPrice': row[20],
                'Currency': row[21],
                'Customer_prf_Number': row[22],
                'Vendors_Brands': row[23],
                'Category': row[24],
                'Partnumber': row[25],
                'Qty': row[26],
                'PackId': row[27],
                'BoxId': row[28],
                'Serilized_NonSerilized': row[29],
                'SerialNumber': row[30],
                'UnitPrice': row[31],
                'TotalAmount': row[32],
                'VendorInvNo': row[33],
                'ECI_Date': row[34],
                'DueDate_ExpiryDate': row[35],
                'ASP_FDR_Date': row[36],
                'Call_Id': row[37],
                'Detailed_Issue_Product': row[38],
                'PRF_Num': row[39],
                'RR_Num': row[40],
                'RR_Date': row[41],
                'Unit_Collection_Date': row[42],
                'CM_Num': row[43],
                'CM_Amount': row[44],
                'RPO_Num': row[45],
                'RPO_Date': row[46],
                'GRN_Date': row[47],
                'DHL_AWB_Num': row[48],
                'Forwarder_Code': row[49],
                'Claims_Status': row[50],
                'AdditionalClaimsstatus': row[51],
                'Remarks': row[52],
                'Status': row[53],
                'CN_No': row[54],
                'CN_Date': row[55],
                'CN_Value': row[56],
                'CA': row[57],
                'Ageingdays': row[58],
            }

            # Remove 'id' key to avoid using it in update_or_create
            claim_id = claim_data.pop('id')

            # Check if the record exists, and update or create
            ClaimsTracker.objects.update_or_create(id=claim_id, defaults=claim_data)


    Departments = Department.objects.all()
    Divisionss = Divisions.objects.all()
    ClaimTypes = ClaimType.objects.all()
    Regionss = Regions.objects.all()
    Vendors_Brandss = Vendors_Brands.objects.all()
    Categorys = Category.objects.all()
    ForwarderCodes = Forwarder_Code.objects.all()
    ClaimsStatuss = ClaimsStatus.objects.all()
    AdditionalClaimsStatuss = AdditionalClaimsStatus.objects.all()
    Statuss = Status.objects.all()
    ClaimsTrackers = ClaimsTracker.objects.all()
    condition = True

    Context = {'condition': condition, 'username': request.session['username'], 'Statuss': Statuss,
               'AdditionalClaimsStatuss': AdditionalClaimsStatuss, 'ClaimsStatuss': ClaimsStatuss,
               'ForwarderCodes': ForwarderCodes, 'Categorys': Categorys, 'Vendors_Brandss': Vendors_Brandss,
               'ClaimsTrackers': ClaimsTrackers, 'Departments': Departments, 'Divisionss': Divisionss,
               'ClaimTypes': ClaimTypes, 'Regionss': Regionss}

    return render(request, 'index.html', context=Context)

def export_claims_tracker(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Claims Tracker"

    # Define the column headers
    columns = [
        'id', 'Department', 'Divisions', 'ClaimType', 'Regions', 'RequestFrom',
        'Originator', 'ClaimedDate', 'ClaimRefNumber', 'Ponumber', 'cbn', 'hporder',
        'customer_request_date', 'Partner_Customername', 'CustomerCode',
        'Partner_location', 'CustomerReference', 'RedingtonInvoice',
        'Redington_Invoice_Date', 'SellUnitPrice', 'TotalSellPrice', 'Currency',
        'Customer_prf_Number', 'Vendors_Brands', 'Category', 'Partnumber', 'Qty',
        'PackId', 'BoxId', 'Serilized_NonSerilized', 'SerialNumber', 'UnitPrice',
        'TotalAmount', 'VendorInvNo', 'ECI_Date', 'DueDate_ExpiryDate', 'ASP_FDR_Date',
        'Call_Id', 'Detailed_Issue_Product', 'PRF_Num', 'RR_Num', 'RR_Date',
        'Unit_Collection_Date', 'CM_Num', 'CM_Amount', 'RPO_Num', 'RPO_Date',
        'GRN_Date', 'DHL_AWB_Num', 'Forwarder_Code', 'Claims_Status',
        'AdditionalClaimsstatus', 'Remarks', 'Status', 'CN_No', 'CN_Date', 'CN_Value',
        'CA', 'Ageingdays'
    ]

    # Write the column headers to the first row
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Create a date style
    date_style = NamedStyle(name="date_style", number_format="DD-MM-YYYY")

    # Query the database and write the data rows
    for row_num, claim in enumerate(ClaimsTracker.objects.all(), 2):
        ws.cell(row=row_num, column=1, value=claim.id)
        ws.cell(row=row_num, column=2, value=claim.Department)
        ws.cell(row=row_num, column=3, value=claim.Divisions)
        ws.cell(row=row_num, column=4, value=claim.ClaimType)
        ws.cell(row=row_num, column=5, value=claim.Regions)
        ws.cell(row=row_num, column=6, value=claim.RequestFrom)
        ws.cell(row=row_num, column=7, value=claim.Originator)
        ws.cell(row=row_num, column=8, value=make_naive(
            claim.ClaimedDate) if claim.ClaimedDate else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=9, value=claim.ClaimRefNumber)
        ws.cell(row=row_num, column=10, value=claim.Ponumber)
        ws.cell(row=row_num, column=11, value=claim.cbn)
        ws.cell(row=row_num, column=12, value=claim.hporder)
        ws.cell(row=row_num, column=13, value=make_naive(
            claim.customer_request_date) if claim.customer_request_date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=14, value=claim.Partner_Customername)
        ws.cell(row=row_num, column=15, value=claim.CustomerCode)
        ws.cell(row=row_num, column=16, value=claim.Partner_location)
        ws.cell(row=row_num, column=17, value=claim.CustomerReference)
        ws.cell(row=row_num, column=18, value=claim.RedingtonInvoice)
        ws.cell(row=row_num, column=19, value=make_naive(
            claim.Redington_Invoice_Date) if claim.Redington_Invoice_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=20, value=claim.SellUnitPrice)
        ws.cell(row=row_num, column=21, value=claim.TotalSellPrice)
        ws.cell(row=row_num, column=22, value=claim.Currency)
        ws.cell(row=row_num, column=23, value=claim.Customer_prf_Number)
        ws.cell(row=row_num, column=24, value=claim.Vendors_Brands)
        ws.cell(row=row_num, column=25, value=claim.Category)
        ws.cell(row=row_num, column=26, value=claim.Partnumber)
        ws.cell(row=row_num, column=27, value=claim.Qty)
        ws.cell(row=row_num, column=28, value=claim.PackId)
        ws.cell(row=row_num, column=29, value=claim.BoxId)
        ws.cell(row=row_num, column=30, value=claim.Serilized_NonSerilized)
        ws.cell(row=row_num, column=31, value=claim.SerialNumber)
        ws.cell(row=row_num, column=32, value=claim.UnitPrice)
        ws.cell(row=row_num, column=33, value=claim.TotalAmount)
        ws.cell(row=row_num, column=34, value=claim.VendorInvNo)
        ws.cell(row=row_num, column=35,
                value=make_naive(claim.ECI_Date) if claim.ECI_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=36, value=make_naive(
            claim.DueDate_ExpiryDate) if claim.DueDate_ExpiryDate else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=37, value=make_naive(
            claim.ASP_FDR_Date) if claim.ASP_FDR_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=38, value=claim.Call_Id)
        ws.cell(row=row_num, column=39, value=claim.Detailed_Issue_Product)
        ws.cell(row=row_num, column=40, value=claim.PRF_Num)
        ws.cell(row=row_num, column=41, value=claim.RR_Num)
        ws.cell(row=row_num, column=42,
                value=make_naive(claim.RR_Date) if claim.RR_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=43, value=make_naive(
            claim.Unit_Collection_Date) if claim.Unit_Collection_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=44, value=claim.CM_Num)
        ws.cell(row=row_num, column=45, value=claim.CM_Amount)
        ws.cell(row=row_num, column=46, value=claim.RPO_Num)
        ws.cell(row=row_num, column=47,
                value=make_naive(claim.RPO_Date) if claim.RPO_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=48,
                value=make_naive(claim.GRN_Date) if claim.GRN_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=49, value=claim.DHL_AWB_Num)
        ws.cell(row=row_num, column=50, value=claim.Forwarder_Code)
        ws.cell(row=row_num, column=51, value=claim.Claims_Status)
        ws.cell(row=row_num, column=52, value=claim.AdditionalClaimsstatus)
        ws.cell(row=row_num, column=53, value=claim.Remarks)
        ws.cell(row=row_num, column=54, value=claim.Status)
        ws.cell(row=row_num, column=55, value=claim.CN_No)
        ws.cell(row=row_num, column=56,
                value=make_naive(claim.CN_Date) if claim.CN_Date else None).number_format = date_style.number_format
        ws.cell(row=row_num, column=57, value=claim.CN_Value)
        ws.cell(row=row_num, column=58, value=claim.CA)
        ws.cell(row=row_num, column=59, value=claim.Ageingdays)

    # Set the response content type and headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=claims_tracker.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

def claimformupdate(request):

    hiddenid = request.POST['hiddenid']

    department = request.POST['department']
    division = request.POST['division']
    claimtype = request.POST['claimtype']
    regions = request.POST['regions']
    requestfrom = request.POST['requestfrom']
    originator = request.POST['originator']

    if request.POST.get("claimdate"):
        claimdate = datetime.strptime(request.POST.get("claimdate"), "%d-%m-%Y").date()
    else:
        claimdate = None

    ClaimRef = request.POST['ClaimRef']
    ponum = request.POST['ponum']
    CBN = request.POST['CBN']
    hporder = request.POST['hporder']

    if request.POST.get("customerreqdate"):
        customerreqdate = datetime.strptime(request.POST.get("customerreqdate"), "%d-%m-%Y").date()
    else:
        customerreqdate = None

    PartnerCustomername = request.POST['PartnerCustomername']
    CustomerCode = request.POST['CustomerCode']
    Partnerlocation = request.POST['Partnerlocation']
    Customerreference = request.POST['Customerreference']
    Redingtoninvoice = request.POST['Redingtoninvoice']

    if request.POST.get("Redingtoninvoicedate"):
        Redingtoninvoicedate = datetime.strptime(request.POST.get("Redingtoninvoicedate"), "%d-%m-%Y").date()
    else:
        Redingtoninvoicedate = None

    SellunitPrice = request.POST['SellunitPrice']
    TotalSellPrice = request.POST['TotalSellPrice']
    Currenc = request.POST['Currency']
    CustomerPRFNo = request.POST['CustomerPRFNo']
    vendors = request.POST['vendors']
    Category_new = request.POST['Category']
    Part = request.POST['Part']
    Qty = request.POST['Qty']
    PackID = request.POST['PackID']
    BoxID = request.POST['BoxID']
    SerilizedNonSerilized = request.POST['SerilizedNonSerilized']
    Serial = request.POST['Serial']
    unitprice = request.POST['unitprice']
    TotalAmount = request.POST['TotalAmount$']
    VendorInvNo = request.POST['VendorInvNo']

    if request.POST.get("ecidatenew"):
        ecidatenew = datetime.strptime(request.POST.get("ecidatenew"), "%d-%m-%Y").date()
    else:
        ecidatenew = None

    if request.POST.get("Duedate"):
        Duedate = datetime.strptime(request.POST.get("Duedate"), "%d-%m-%Y").date()
    else:
        Duedate = None

    if request.POST.get("aspfdr"):
        aspfdr = datetime.strptime(request.POST.get("aspfdr"), "%d-%m-%Y").date()
    else:
        aspfdr = None

    CallID = request.POST['CallID']
    detailed = request.POST['detailed']
    PRFNo = request.POST['PRFNo']
    RR = request.POST['RR']

    if request.POST.get("RRdate"):
        RRdate = datetime.strptime(request.POST.get("RRdate"), "%d-%m-%Y").date()
    else:
        RRdate = None

    if request.POST.get("unitcollection"):
        unitcollection = datetime.strptime(request.POST.get("unitcollection"), "%d-%m-%Y").date()
    else:
        unitcollection = None

    CM = request.POST['CM']
    CMAMOUNT = request.POST['CMAMOUNT']
    RPO = request.POST['RPO']

    if request.POST.get("RPODATE"):
        RPODATE = datetime.strptime(request.POST.get("RPODATE"), "%d-%m-%Y").date()
    else:
        RPODATE = None

    if request.POST.get("grndate"):
        grndate = datetime.strptime(request.POST.get("grndate"), "%d-%m-%Y").date()
    else:
        grndate = None

    dhl = request.POST['dhl']
    FORWARDERCODE = request.POST['FORWARDERCODE']
    ClaimsStatus_new = request.POST['ClaimsStatus']
    AdditionalClaimsStatus_new = request.POST['AdditionalClaimsStatus']
    Remarks = request.POST['Remarks']
    status_new = request.POST['status']
    cnno = request.POST['cnno']

    if request.POST.get("cndate"):
        cndate = datetime.strptime(request.POST.get("cndate"), "%d-%m-%Y").date()
    else:
        cndate = None

    cnvalue = request.POST['cnvalue']
    ca = request.POST['ca']
    ageing = request.POST['ageing']

    existing_claim_tracker = ClaimsTracker.objects.get(id=hiddenid)

    # Update the fields
    existing_claim_tracker.Department = department
    existing_claim_tracker.Divisions = division
    existing_claim_tracker.ClaimType = claimtype
    existing_claim_tracker.Regions = regions
    existing_claim_tracker.RequestFrom = requestfrom
    existing_claim_tracker.Originator = originator
    existing_claim_tracker.ClaimedDate = claimdate
    existing_claim_tracker.Ponumber = ponum
    existing_claim_tracker.ClaimRefNumber = ClaimRef
    existing_claim_tracker.cbn = CBN
    existing_claim_tracker.hporder = hporder
    existing_claim_tracker.customer_request_date = customerreqdate
    existing_claim_tracker.Partner_Customername = PartnerCustomername
    existing_claim_tracker.CustomerCode = CustomerCode
    existing_claim_tracker.Partner_location = Partnerlocation
    existing_claim_tracker.CustomerReference = Customerreference
    existing_claim_tracker.RedingtonInvoice = Redingtoninvoice
    existing_claim_tracker.Redington_Invoice_Date = Redingtoninvoicedate
    existing_claim_tracker.SellUnitPrice = SellunitPrice
    existing_claim_tracker.TotalSellPrice = TotalSellPrice
    existing_claim_tracker.Currency = Currenc
    existing_claim_tracker.Customer_prf_Number = CustomerPRFNo
    existing_claim_tracker.Vendors_Brands = vendors
    existing_claim_tracker.Category = Category_new
    existing_claim_tracker.Partnumber = Part
    existing_claim_tracker.Qty = Qty
    existing_claim_tracker.PackId = PackID
    existing_claim_tracker.BoxId = BoxID
    existing_claim_tracker.Serilized_NonSerilized = SerilizedNonSerilized
    existing_claim_tracker.SerialNumber = Serial
    existing_claim_tracker.UnitPrice = unitprice
    existing_claim_tracker.TotalAmount = TotalAmount
    existing_claim_tracker.VendorInvNo = VendorInvNo
    existing_claim_tracker.ECI_Date = ecidatenew
    existing_claim_tracker.DueDate_ExpiryDate = Duedate
    existing_claim_tracker.ASP_FDR_Date = aspfdr
    existing_claim_tracker.Call_Id = CallID
    existing_claim_tracker.Detailed_Issue_Product = detailed
    existing_claim_tracker.PRF_Num = PRFNo
    existing_claim_tracker.RR_Num = RR
    existing_claim_tracker.RR_Date = RRdate
    existing_claim_tracker.Unit_Collection_Date = unitcollection
    existing_claim_tracker.CM_Num = CM
    existing_claim_tracker.CM_Amount = CMAMOUNT
    existing_claim_tracker.RPO_Num = RPO
    existing_claim_tracker.RPO_Date = RPODATE
    existing_claim_tracker.GRN_Date = grndate
    existing_claim_tracker.DHL_AWB_Num = dhl
    existing_claim_tracker.Forwarder_Code = FORWARDERCODE
    existing_claim_tracker.Claims_Status = ClaimsStatus_new
    existing_claim_tracker.AdditionalClaimsstatus = AdditionalClaimsStatus_new
    existing_claim_tracker.Remarks = Remarks
    existing_claim_tracker.Status = status_new
    existing_claim_tracker.CN_No = cnno
    existing_claim_tracker.CN_Date = cndate
    existing_claim_tracker.CN_Value = cnvalue
    existing_claim_tracker.CA = ca
    existing_claim_tracker.Ageingdays = ageing

    # Save the updated record
    existing_claim_tracker.save()

    Departments = Department.objects.all()
    Divisionss = Divisions.objects.all()
    ClaimTypes = ClaimType.objects.all()
    Regionss = Regions.objects.all()
    Vendors_Brandss = Vendors_Brands.objects.all()
    Categories = Category.objects.all()
    ForwarderCodes = Forwarder_Code.objects.all()
    ClaimsStatuss = ClaimsStatus.objects.all()
    AdditionalClaimsStatuss = AdditionalClaimsStatus.objects.all()
    Statuss = Status.objects.all()
    ClaimsTrackers = ClaimsTracker.objects.all()
    Currencys = Currency.objects.all()
    Serilizeds = Serilized_NonSerilized.objects.all()

    condition = False

    Context = {'Currencys': Currencys ,'Serilizeds':Serilizeds,'condition': condition, 'username': request.session['username'], 'Statuss': Statuss,
               'AdditionalClaimsStatuss': AdditionalClaimsStatuss, 'ClaimsStatuss': ClaimsStatuss,
               'ForwarderCodes': ForwarderCodes, 'Categorys': Categories, 'Vendors_Brandss': Vendors_Brandss,
               'ClaimsTrackers': ClaimsTrackers, 'Departments': Departments, 'Divisionss': Divisionss,
               'ClaimTypes': ClaimTypes, 'Regionss': Regionss}

    return render(request, 'index.html', context=Context)


def edit(request,id):

    Departments = Department.objects.all()
    Divisionss = Divisions.objects.all()
    ClaimTypes = ClaimType.objects.all()
    Regionss = Regions.objects.all()
    Vendors_Brandss = Vendors_Brands.objects.all()
    Categorys = Category.objects.all()
    ForwarderCodes = Forwarder_Code.objects.all()
    ClaimsStatuss = ClaimsStatus.objects.all()
    AdditionalClaimsStatuss = AdditionalClaimsStatus.objects.all()
    Statuss = Status.objects.all()
    Claimstracker = get_object_or_404(ClaimsTracker, id=id)
    Currencys = Currency.objects.all()

    if str(Claimstracker.ClaimedDate)=="None":
        formatted_ClaimedDate =""
    else:
        ClaimedDate = timezone.localtime(Claimstracker.ClaimedDate)
        formatted_ClaimedDate = ClaimedDate.strftime("%d-%m-%Y")

    if str(Claimstracker.customer_request_date)=="None":
        formatted_customerrequestdate =""
    else:
        customerrequestdate = timezone.localtime(Claimstracker.customer_request_date)
        formatted_customerrequestdate = customerrequestdate.strftime("%d-%m-%Y")

    if str(Claimstracker.Redington_Invoice_Date)=="None":
        formatted_RedingtonInvoiceDate =""
    else:
        RedingtonInvoiceDate = timezone.localtime(Claimstracker.Redington_Invoice_Date)
        formatted_RedingtonInvoiceDate = RedingtonInvoiceDate.strftime("%d-%m-%Y")


    if str(Claimstracker.ECI_Date)=="None":
        formatted_ecidate =""
    else:
        ecidate = timezone.localtime(Claimstracker.ECI_Date)
        formatted_ecidate = ecidate.strftime("%d-%m-%Y")

    if str(Claimstracker.DueDate_ExpiryDate)=="None":
        formatted_duedateexpirydate =""
    else:
        duedateexpirydate = timezone.localtime(Claimstracker.DueDate_ExpiryDate)
        formatted_duedateexpirydate = duedateexpirydate.strftime("%d-%m-%Y")

    if str(Claimstracker.ASP_FDR_Date)=="None":
        formatted_aspfdrdate =""
    else:
        aspfdrdate = timezone.localtime(Claimstracker.ASP_FDR_Date)
        formatted_aspfdrdate = aspfdrdate.strftime("%d-%m-%Y")


    if str(Claimstracker.RR_Date) == "None":
        formatted_RR_Date = ""
    else:
        RR_Date = timezone.localtime(Claimstracker.RR_Date)
        formatted_RR_Date = RR_Date.strftime("%d-%m-%Y")


    if str(Claimstracker.Unit_Collection_Date) == "None":
        formatted_Unit_Collection_Date = ""
    else:
        Unit_Collection_Date = timezone.localtime(Claimstracker.Unit_Collection_Date)
        formatted_Unit_Collection_Date = Unit_Collection_Date.strftime("%d-%m-%Y")


    if str(Claimstracker.RPO_Date) == "None":
        formatted_RPO_Date = ""
    else:
        RPO_Date = timezone.localtime(Claimstracker.RPO_Date)
        formatted_RPO_Date = RPO_Date.strftime("%d-%m-%Y")


    if str(Claimstracker.GRN_Date) == "None":
        formatted_GRN_Date = ""
    else:
        GRN_Date = timezone.localtime(Claimstracker.GRN_Date)
        formatted_GRN_Date = GRN_Date.strftime("%d-%m-%Y")


    if str(Claimstracker.CN_Date) == "None":
        formatted_CN_Date = ""
    else:
        CN_Date = timezone.localtime(Claimstracker.CN_Date)
        formatted_CN_Date = CN_Date.strftime("%d-%m-%Y")

    Currencys = Currency.objects.all()
    Serilizeds = Serilized_NonSerilized.objects.all()


    Context = {'Currencys':Currencys,'Serilizeds':Serilizeds, 'id':id,'Currencys':Currencys, 'formatted_CN_Date':formatted_CN_Date, 'formatted_GRN_Date':formatted_GRN_Date, 'formatted_RPO_Date':formatted_RPO_Date, 'formatted_Unit_Collection_Date':formatted_Unit_Collection_Date,'formatted_RR_Date':formatted_RR_Date,'formatted_aspfdrdate':formatted_aspfdrdate,'formatted_duedateexpirydate':formatted_duedateexpirydate,'formatted_ecidate':formatted_ecidate,'formatted_customerrequestdate':formatted_customerrequestdate, 'formatted_RedingtonInvoiceDate':formatted_RedingtonInvoiceDate, 'formatted_ClaimedDate':formatted_ClaimedDate, 'username': request.session['username'], 'Statuss': Statuss,
               'AdditionalClaimsStatuss': AdditionalClaimsStatuss, 'ClaimsStatuss': ClaimsStatuss,
               'ForwarderCodes': ForwarderCodes, 'Categorys': Categorys, 'Vendors_Brandss': Vendors_Brandss,
               'ClaimsTracker': Claimstracker, 'Departments': Departments, 'Divisionss': Divisionss,
               'ClaimTypes': ClaimTypes, 'Regionss': Regionss}

    return render(request, 'indexedit.html', context=Context)


def delete(request,id):

    Claimstracker = get_object_or_404(ClaimsTracker, id=id)
    Claimstracker.delete()

    Departments = Department.objects.all()
    Divisionss = Divisions.objects.all()
    ClaimTypes = ClaimType.objects.all()
    Regionss = Regions.objects.all()
    Vendors_Brandss = Vendors_Brands.objects.all()
    Categorys = Category.objects.all()
    ForwarderCodes = Forwarder_Code.objects.all()
    ClaimsStatuss = ClaimsStatus.objects.all()
    AdditionalClaimsStatuss = AdditionalClaimsStatus.objects.all()
    Statuss = Status.objects.all()
    ClaimsTrackers = ClaimsTracker.objects.all()
    Currencys = Currency.objects.all()
    Serilizeds = Serilized_NonSerilized.objects.all()

    condition=True

    Context = {'Serilizeds':Serilizeds,'Currencys':Currencys, 'condition':condition, 'username': request.session['username'], 'Statuss': Statuss,
               'AdditionalClaimsStatuss': AdditionalClaimsStatuss, 'ClaimsStatuss': ClaimsStatuss,
               'ForwarderCodes': ForwarderCodes, 'Categorys': Categorys, 'Vendors_Brandss': Vendors_Brandss,
               'ClaimsTrackers': ClaimsTrackers, 'Departments': Departments, 'Divisionss': Divisionss,
               'ClaimTypes': ClaimTypes, 'Regionss': Regionss}

    return render(request, 'index.html', context=Context)

def home(request):
    if 'username' in request.session:
        Departments = Department.objects.all()
        Divisionss = Divisions.objects.all()
        ClaimTypes = ClaimType.objects.all()
        Regionss = Regions.objects.all()
        Vendors_Brandss = Vendors_Brands.objects.all()
        Categorys = Category.objects.all()
        ForwarderCodes = Forwarder_Code.objects.all()
        ClaimsStatuss = ClaimsStatus.objects.all()
        AdditionalClaimsStatuss = AdditionalClaimsStatus.objects.all()
        Statuss = Status.objects.all()
        ClaimsTrackers=ClaimsTracker.objects.all()
        Currencys = Currency.objects.all()
        Serilizeds = Serilized_NonSerilized.objects.all()

        Context = {'Serilizeds':Serilizeds,'Currencys':Currencys, 'username': request.session['username'], 'Statuss': Statuss,
                   'AdditionalClaimsStatuss': AdditionalClaimsStatuss, 'ClaimsStatuss': ClaimsStatuss,
                   'ForwarderCodes': ForwarderCodes, 'Categorys': Categorys, 'Vendors_Brandss': Vendors_Brandss,
                    'ClaimsTrackers':ClaimsTrackers ,'Departments': Departments, 'Divisionss': Divisionss, 'ClaimTypes': ClaimTypes, 'Regionss': Regionss}

        return render(request, 'index.html', context=Context)
    else:
        usernames = Login.objects.all()
        context = {'usernames': usernames}
        return render(request, 'login.html', context=context)


def login(request):
    usernames = Login.objects.all()
    context = {'usernames': usernames}
    return render(request, 'login.html', context=context)

def logout(request):
    del request.session['username']
    usernames = Login.objects.all()
    context = {'usernames': usernames}
    return render(request, 'login.html',context=context)

def processlogin(request):
    email = request.POST["Email"]
    request.session['email'] = email
    password = request.POST["Password"]
    login_query = Login.objects.all()
    loginbool = False

    logger.info(str(email) + str(" loggedin ") + str(datetime.now()))

    for login in login_query:
        if str(login.username).upper() == str(email).upper() and str(login.password).upper() == str(password).upper():
            loginbool = True

    if loginbool == True:
        request.session['username'] = str(email).split("@")[0].split(".")[0]
        Departments = Department.objects.all()
        Divisionss = Divisions.objects.all()
        ClaimTypes= ClaimType.objects.all()
        Regionss = Regions.objects.all()
        Vendors_Brandss= Vendors_Brands.objects.all()
        Categorys=Category.objects.all()
        ForwarderCodes=Forwarder_Code.objects.all()
        ClaimsStatuss=ClaimsStatus.objects.all()
        AdditionalClaimsStatuss= AdditionalClaimsStatus.objects.all()
        Statuss= Status.objects.all()
        ClaimsTrackers = ClaimsTracker.objects.all()
        Currencys = Currency.objects.all()
        Serilizeds = Serilized_NonSerilized.objects.all()


        Context = {'Serilizeds':Serilizeds, 'Currencys':Currencys,'username': request.session['username'],'ClaimsTrackers':ClaimsTrackers,'Statuss':Statuss,'AdditionalClaimsStatuss':AdditionalClaimsStatuss,'ClaimsStatuss':ClaimsStatuss,'ForwarderCodes':ForwarderCodes,'Categorys':Categorys,'Vendors_Brandss':Vendors_Brandss,'Departments':Departments,'Divisionss':Divisionss,'ClaimTypes':ClaimTypes,'Regionss':Regionss}
        return render(request, 'index.html', context=Context)
    else:
        usernames = Login.objects.all()
        context = {'usernames': usernames}
        return render(request, 'login.html', context=context)

def claimformsubmit(request):
    if 'username' in request.session:

        if ClaimsTracker.objects.exists():
            ids = ClaimsTracker.objects.values_list('id', flat=True)
            last_segments = [int(claim_id.split('-')[-1]) for claim_id in ids]
            max_value = max(last_segments)+1
            max_value_formatted = f"{max_value:06}"
            id = str("REDCLMFY24-25-")+str(max_value_formatted)
        else:
            id = "REDCLMFY24-25-000001"

        department = request.POST['department']
        division = request.POST['division']
        claimtype= request.POST['claimtype']
        regions=request.POST['regions']
        requestfrom= request.POST['requestfrom']
        originator = request.POST['originator']

        if request.POST.get("claimdate"):
            claimdate = datetime.strptime(request.POST.get("claimdate"), "%d-%m-%Y").date()
        else:
            claimdate = None

        ClaimRef = request.POST['ClaimRef']
        ponum = request.POST['ponum']
        CBN= request.POST['CBN']
        hporder=request.POST['hporder']

        if request.POST.get("customerreqdate"):
            customerreqdate = datetime.strptime(request.POST.get("customerreqdate"), "%d-%m-%Y").date()
        else:
            customerreqdate = None

        PartnerCustomername=request.POST['PartnerCustomername']
        CustomerCode= request.POST['CustomerCode']
        Partnerlocation=request.POST['Partnerlocation']
        Customerreference=request.POST['Customerreference']
        Redingtoninvoice=request.POST['Redingtoninvoice']

        if request.POST.get("Redinvdate"):
            Redingtoninvoicedate = datetime.strptime(request.POST.get("Redinvdate"), "%d-%m-%Y").date()
        else:
            Redingtoninvoicedate = None

        SellunitPrice= request.POST['SellunitPrice']
        TotalSellPrice= request.POST['TotalSellPrice']
        Currency=request.POST['Currency']
        CustomerPRFNo=request.POST['CustomerPRFNo']
        vendors = request.POST['vendors']
        Category=request.POST['Category']
        Part= request.POST['Part']
        Qty= request.POST['Qty']
        PackID = request.POST['PackID']
        BoxID = request.POST['BoxID']
        SerilizedNonSerilized = request.POST['SerilizedNonSerilized']
        Serial = request.POST['Serial']
        unitprice = request.POST['unitprice']
        TotalAmount = request.POST['TotalAmount$']
        VendorInvNo = request.POST['VendorInvNo']

        if request.POST.get("ecidatenew"):
            ecidatenew = datetime.strptime(request.POST.get("ecidatenew"), "%d-%m-%Y").date()
        else:
            ecidatenew = None

        if request.POST.get("Duedate"):
            Duedate = datetime.strptime(request.POST.get("Duedate"), "%d-%m-%Y").date()
        else:
            Duedate = None

        if request.POST.get("aspfdr"):
            aspfdr = datetime.strptime(request.POST.get("aspfdr"), "%d-%m-%Y").date()
        else:
            aspfdr = None

        CallID = request.POST['CallID']
        detailed= request.POST['detailed']
        PRFNo = request.POST['PRFNo']
        RR = request.POST['RR']

        if request.POST.get("RRdate"):
            RRdate = datetime.strptime(request.POST.get("RRdate"), "%d-%m-%Y").date()
        else:
            RRdate = None

        if request.POST.get("unitcollection"):
            unitcollection = datetime.strptime(request.POST.get("unitcollection"), "%d-%m-%Y").date()
        else:
            unitcollection = None


        CM  = request.POST['CM']
        CMAMOUNT = request.POST['CMAMOUNT']
        RPO=request.POST['RPO']

        if request.POST.get("RPODATE"):
            RPODATE = datetime.strptime(request.POST.get("RPODATE"), "%d-%m-%Y").date()
        else:
            RPODATE = None


        if request.POST.get("grndate"):
            grndate = datetime.strptime(request.POST.get("grndate"), "%d-%m-%Y").date()
        else:
            grndate = None


        dhl=request.POST['dhl']
        FORWARDERCODE = request.POST['FORWARDERCODE']
        ClaimsStatus=request.POST['ClaimsStatus']
        AdditionalClaimsStatus = request.POST['AdditionalClaimsStatus']
        Remarks = request.POST['Remarks']
        status = request.POST['status']
        cnno = request.POST['cnno']

        if request.POST.get("cndate"):
            cndate = datetime.strptime(request.POST.get("cndate"), "%d-%m-%Y").date()
        else:
            cndate = None

        cnvalue = request.POST['cnvalue']
        ca = request.POST['ca']
        ageing= request.POST['ageing']

        claim_tracker = ClaimsTracker(
            id=id,
            Department=department,
            Divisions=division,
            ClaimType=claimtype,
            Regions =regions,
            RequestFrom=requestfrom,
            Originator=originator,
            ClaimedDate=claimdate,
            ClaimRefNumber=ClaimRef,
            Ponumber=ponum,
            cbn=CBN,
            hporder=hporder,
            customer_request_date =customerreqdate,
            Partner_Customername =PartnerCustomername,
            CustomerCode=CustomerCode,
            Partner_location=Partnerlocation,
            CustomerReference=Customerreference,
            RedingtonInvoice=Redingtoninvoice,
            Redington_Invoice_Date=Redingtoninvoicedate,
            SellUnitPrice=SellunitPrice,
            TotalSellPrice=TotalSellPrice,
            Currency=Currency,
            Customer_prf_Number =CustomerPRFNo,
            Vendors_Brands =vendors,
            Category=Category,
            Partnumber =Part,
            Qty=Qty,
            PackId=PackID,
            BoxId =BoxID,
            Serilized_NonSerilized=SerilizedNonSerilized,
            SerialNumber =Serial,
            UnitPrice=unitprice,
            TotalAmount=TotalAmount,
            VendorInvNo=VendorInvNo,
            ECI_Date=ecidatenew,
            DueDate_ExpiryDate=Duedate,
            ASP_FDR_Date=aspfdr,
            Call_Id=CallID,
            Detailed_Issue_Product =detailed,
            PRF_Num =PRFNo,
            RR_Num=RR,
            RR_Date=RRdate,
            Unit_Collection_Date=unitcollection,
            CM_Num=CM,
            CM_Amount =CMAMOUNT,
            RPO_Num =RPO,
            RPO_Date=RPODATE,
            GRN_Date=grndate,
            DHL_AWB_Num=dhl,
            Forwarder_Code=FORWARDERCODE,
            Claims_Status =ClaimsStatus,
            AdditionalClaimsstatus=AdditionalClaimsStatus,
            Remarks=Remarks,
            Status=status,
            CN_No=cnno,
            CN_Date=cndate,
            CN_Value =cnvalue,
            CA=ca,
            Ageingdays=ageing
        )

        claim_tracker.full_clean()
        claim_tracker.save()
        return redirect('physicalclaims:home')






























        return HttpResponse(department)
    else:
        usernames = Login.objects.all()
        context = {'usernames': usernames}
        return render(request, 'login.html', context=context)
